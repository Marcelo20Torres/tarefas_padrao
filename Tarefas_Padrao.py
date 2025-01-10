import customtkinter as ctk
import tkinter as tk
from openpyxl import load_workbook
import pandas as pd
from CTkMessagebox import CTkMessagebox
from tkinter import messagebox
from pathlib import Path
from PIL import Image


def formatar_cnpj(event=None):
    cnpj_ = CNPJ.get()

    cnpj_numeros = ''.join(filter(str.isdigit, cnpj_))

    if len(cnpj_numeros) <= 2:
        cnpj_formatado = cnpj_numeros
    elif len(cnpj_numeros) <= 5:
        cnpj_formatado = f"{cnpj_numeros[:2]}.{cnpj_numeros[2:]}"
    elif len(cnpj_numeros) <= 8:
        cnpj_formatado = f"{cnpj_numeros[:2]}.{cnpj_numeros[2:5]}.{cnpj_numeros[5:]}"
    elif len(cnpj_numeros) <= 12:
        cnpj_formatado = f"{cnpj_numeros[:2]}.{cnpj_numeros[2:5]}.{cnpj_numeros[5:8]}/{cnpj_numeros[8:]}"
    else:
        cnpj_formatado = f"{cnpj_numeros[:2]}.{cnpj_numeros[2:5]}.{cnpj_numeros[5:8]}/{cnpj_numeros[8:12]}-{cnpj_numeros[12:14]}"

    CNPJ.delete(0, tk.END)
    CNPJ.insert(0, cnpj_formatado)

def filtrar_planilha():
    localidade = caixa_localidade.get()
    Aba1 = caixa_area.get()
    regime = caixa_regime.get()
    tipo = caixa_tipo.get()
    ramo = caixa_ramo.get()
    cnpj = CNPJ.get()

    if not localidade.strip() or not Aba1.strip() or not regime.strip() or not tipo.strip() or not ramo.strip() or not cnpj.strip():
        messagebox.showwarning("Aviso","Todos os campos devem estar preenchidos!")
        return

    if Aba1 == "Contábil/Fiscal - Online":
        Aba = "Online"
    elif Aba1 == "Contábil/Fiscal - PEGN":
        Aba = "Simples Nacional"
    elif Aba1 == "Contábil/Fiscal - Digital":
        Aba = "Digital"
    
    if regime == "Simples Nacional (LC Presum":
        Aba = "Simples Nacional (LC Presum"
        
    caminho = Path(__file__).parent / 'TAREFAS PADRÃO NUUBES IND COM E SERVICOS_v65_04_04_2024.xlsx'
        
    df = pd.read_excel(caminho, sheet_name=f'{Aba}', header=1)
        
    #Empresa de Goiânia ?
    if localidade == "Sim":
        df_filtrado = df[df['Empresa de Goiânia?'] == f'{localidade}']
    elif localidade == "Não":
        df_filtrado = df[df['Empresa demais municipios?'] == 'Sim']
        
    #Regime Tributario
    if regime == "Simples Nacional (LC Presum":
        df_filtrado = df_filtrado[(df_filtrado['Regime tributário'] == 'Lucro Presumido') ]
    else:
        df_filtrado = df_filtrado[(df_filtrado['Regime tributário'] == f'{regime}') ]

    #Matriz ou Filial ?
    if tipo == "Matriz" and Aba == "Simples Nacional":
        df_filtrado = df_filtrado[df_filtrado['MATRIZ'] == 'Sim']
    elif tipo == "Matriz" and Aba != "Simples Nacional":
        df_filtrado = df_filtrado[df_filtrado['MATRIZ'] == 'SIM']
    elif tipo == "Filial" and Aba == "Simples Nacional":
        df_filtrado = df_filtrado[df_filtrado['FILIAL'] == 'Sim']
    elif tipo == "Filial" and Aba != "Simples Nacional":
        df_filtrado = df_filtrado[df_filtrado['FILIAL'] == 'SIM']

    #Ramo de Atividade
    if ramo == "Industria":
        df_filtrado = df_filtrado[df_filtrado['Industria?'] == 'SIM']
    elif ramo == "Comércio":
        df_filtrado = df_filtrado[df_filtrado['Comércio?'] == 'SIM']
    elif ramo == "Serviço":
        df_filtrado = df_filtrado[df_filtrado['Serviço?'] == 'SIM']
    elif ramo == "Industria/Comércio":
        df_filtrado = df_filtrado[df_filtrado['INDUSTRIA/COMERCIO'] == 'SIM']
    elif ramo == "Comercio/Serviço":
        df_filtrado = df_filtrado[df_filtrado['COMERCIO/SERVICO'] == 'SIM']
    elif ramo == "Industria/Serviço":
        df_filtrado = df_filtrado[df_filtrado['INDUSTRIA/SERVICO'] == 'SIM']



    resultado = df_filtrado[['Área','Tipo de tarefa']]
    print(resultado)
    caminho_importacao = Path(__file__).parent / 'planilha_tarefas_clientes.xlsx'

    WbI = load_workbook(caminho_importacao)
    sheetI = WbI['Planilha Modelo Tipo de Tarefas']

    dados = resultado.values.tolist()

    linha_inicio = 2
    deletar = 50
    next_row = 0
        
    for row in range(2, deletar + 1):
        for col in sheetI.iter_cols(min_row=row, max_row=row):
            for cell in col:
                cell.value = None

    for i, linha in enumerate(dados, start=linha_inicio):
        for j, valor in enumerate(linha, start=2):
            sheetI.cell(row=i, column=j, value=valor)
            next_row = i

    for row in range(2, next_row + 1):
        sheetI[f'A{row}'] = cnpj
        
    WbI.save(caminho_importacao)

    CTkMessagebox(title="Finalizado", message="Planilha de Tarefas criada com sucesso!", icon="check")



janela = ctk.CTk()
janela.title("Tarefas Padrão")
janela.geometry("600x400")
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

caminho_imagem = Path(__file__).parent / "logo-colorida.png"
imagem_pil = Image.open(caminho_imagem)
imagem = ctk.CTkImage(dark_image=imagem_pil, size=(200, 46))
imagem_simples = ctk.CTkLabel(janela, image=imagem, text="")


caixa_ramo = ctk.CTkComboBox(
    janela,
    values=("", "Industria", "Comércio", "Serviço", "Industria/Comércio", "Comercio/Serviço", "Industria/Serviço")

)

label_ramo = ctk.CTkLabel(
    janela,
    text="Ramo da Empresa:",
    font=("Arial", 15)
)

caixa_tipo = ctk.CTkComboBox(
    janela,
    values=("","Matriz", "Filial"),
    width=140,
    height=29
)

label_tipo = ctk.CTkLabel(
    janela,
    text="Tipo da Empresa:",
    font=("Arial", 15)
)

label_regime = ctk.CTkLabel(
    janela,
    text="Regime Tributário:",
    font=("Arial", 15)

)

caixa_regime = ctk.CTkComboBox(
    janela,
    values=["","Imune Federal", "Isenta do IRPJ", "Lucro Presumido", "Lucro Presumido/RC", "Lucro Real Mensal", "Lucro Real Trimestral", "Simples Nacional - Digital", "Simples Nacional/ICMS Normal", "Simples Nacional", "Simples Nacional/RC", "Simples Nacional (LC Presum"]


)

CNPJ = ctk.CTkEntry(
    janela,
    placeholder_text="CNPJ..."


)

label = ctk.CTkLabel(
    janela,
    text = "Tarefas Padrão",
    font=("Arial",18,"bold")
    
)

botao_fechar = ctk.CTkButton(
    janela,
    text="Fechar", 
    command=janela.quit, 
    width=10, 
    height=30,
    corner_radius=20

    
    )

botao_gerar = ctk.CTkButton(
    janela,
    text="Gerar",
    width=75,
    height=30,
    corner_radius=20,
    command=filtrar_planilha


)

caixa_localidade = ctk.CTkComboBox(

    janela,
    values=["", "Sim", "Não"]
    



)

label_localidade = ctk.CTkLabel(
    janela, 
    text="Empresa de Goiânia: ", 
    font=("Arial", 15)
)

caixa_area = ctk.CTkComboBox(
    janela,
    values=["", "Contábil/Fiscal - PEGN", "Contábil/Fiscal - Digital", "Contábil/Fiscal - Online"]

)

label_area = ctk.CTkLabel(
    janela,
    text="Área da Empresa:",
    font=("Arial", 15)
)


botao_fechar.place(x=415, y=330)
botao_gerar.place(x=500, y=330)
label_localidade.place(x=40, y=110)
caixa_localidade.place(x=40, y=140)
label.place(relx=0.39, rely=0.05)
CNPJ.bind('<KeyRelease>', formatar_cnpj)
CNPJ.place(relx=0.5, rely=0.18, anchor="center")
caixa_regime.place(relx=0.381, y=140)
label_regime.place(relx=0.381, y=110)
caixa_tipo.place(relx=0.7, y=140)
label_tipo.place(relx=0.7, y=110)
caixa_ramo.place(relx=0.20, y=230)
label_ramo.place(relx=0.20, y=200)
caixa_area.place(relx=0.55, y=230)
label_area.place(relx=0.55, y=200)
imagem_simples.place(relx= 0.05, y=315)

janela.mainloop()
